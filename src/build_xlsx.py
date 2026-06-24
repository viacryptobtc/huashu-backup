"""把话术数据列表生成为中文友好的 XLSX。

特性：
- UTF-8，无乱码
- 表头加粗 + 冻结首行 + 自动筛选
- 列宽自适应（中文按 2 字符宽计算）
- 话术正文单元格自动换行、顶端对齐
"""

from __future__ import annotations

import os
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover
    Workbook = None  # type: ignore

HEADERS = ["激活状态", "话术大类", "话术标题", "更新时间", "创建时间", "话术正文"]
FIELDS = ["active", "category", "title", "updated_at", "created_at", "content"]

# 每列最大宽度（字符），正文列不限宽，仅给个起始宽度
MAX_COL_WIDTH = 60


def _display_width(s: str) -> int:
    """估算显示宽度：CJK 字符算 2，其余算 1。"""
    width = 0
    for ch in str(s):
        if "\u4e00" <= ch <= "\u9fff" or "\u3000" <= ch <= "\u303f" \
                or "\uff00" <= ch <= "\uffef":
            width += 2
        else:
            width += 1
    return width


def build_xlsx(rows: list[dict[str, Any]], output_path: str) -> str:
    """生成 XLSX 文件，返回路径。

    rows: fetch_all_macros() 的返回值。
    output_path: 输出 .xlsx 的完整路径。
    """
    if Workbook is None:
        raise RuntimeError("缺少 openpyxl 库，请在虚拟环境中运行: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "话术数据"

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    header_align = Alignment(horizontal="center", vertical="center")

    ws.append(HEADERS)
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # 数据行
    body_align_wrap = Alignment(wrap_text=True, vertical="top")
    for row in rows:
        ws.append([row.get(f, "") for f in FIELDS])

    # 正文列换行 + 顶端对齐
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=len(FIELDS)).alignment = body_align_wrap

    # 冻结首行 + 自动筛选
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"

    # 列宽自适应
    for col_idx, field in enumerate(FIELDS, start=1):
        # 取表头 + 前 200 行采样的最大宽度，避免全量遍历太慢
        sample = [HEADERS[col_idx - 1]] + [
            str(rows[i].get(field, ""))
            for i in range(min(len(rows), 200))
        ]
        width = min(MAX_COL_WIDTH, max(_display_width(s) for s in sample) + 2)
        if field == "content":
            width = 80  # 正文列固定宽一点
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    return output_path


def main() -> int:
    import argparse
    import json
    import sys

    parser = argparse.ArgumentParser(description="把 JSON 数据生成为 XLSX")
    parser.add_argument("--input", "-i", required=True, help="输入 JSON 路径（rows 列表）")
    parser.add_argument("--output", "-o", required=True, help="输出 .xlsx 路径")
    args = parser.parse_args()

    with open(args.input, "r", encoding="utf-8") as f:
        rows = json.load(f)

    build_xlsx(rows, args.output)
    print(f"已生成: {args.output}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    import sys
    sys.exit(main())
